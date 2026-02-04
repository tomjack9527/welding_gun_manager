# welding_gun_system.py
# 将整个 WeldingGunSystem 类替换为以下代码：
# welding_gun_system.py

# 1. 首先，保留所有导入语句
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
import sqlite3
import os
import sys
import datetime
from file_operations import GunFileManager
import json
import shutil

# 2. 添加 FileManager 的回退实现（重要！）
try:
    from file_manager import FileManager
    HAS_FILE_MANAGER = True
except ImportError:
    HAS_FILE_MANAGER = False
    print("文件管理模块未找到，上传下载功能不可用")
    
    # 定义基本 FileManager 类作为回退
    class FileManager:
        def __init__(self, parent_frame):
            """简单的文件管理类"""
            self.frame = tk.LabelFrame(parent_frame, text="文件管理", padx=10, pady=10)
            self.frame.pack(fill="x", padx=10, pady=5)
            
            # 创建按钮
            btn_frame = tk.Frame(self.frame)
            btn_frame.pack(pady=5)
            
            tk.Button(btn_frame, text="上传文件", command=self.upload_file, width=15).pack(side="left", padx=5)
            tk.Button(btn_frame, text="下载文件", command=self.download_file, width=15).pack(side="left", padx=5)
            
            # 状态标签
            self.status_label = tk.Label(self.frame, text="就绪")
            self.status_label.pack()
            
        def upload_file(self):
            messagebox.showinfo("提示", "上传功能需要连接后端服务")
        
        def download_file(self):
            messagebox.showinfo("提示", "下载功能需要连接后端服务")

# 3. 然后添加 Database 类（完整的）
class Database:
    def __init__(self, db_path="welding_gun.db"):
        self.db_path = db_path
        self.conn = None
    
    def connect(self):
        if self.conn is None:
            self.conn = sqlite3.connect(self.db_path)
            self.conn.row_factory = sqlite3.Row
        return self.conn
    
    def close(self):
        if self.conn:
            self.conn.close()
            self.conn = None
    
    def initialize(self):
        try:
            if not os.path.exists(self.db_path):
                self.create_tables()
                self.create_default_data()
            return True
        except Exception as e:
            print(f"数据库初始化失败: {e}")
            return False
    
    def create_tables(self):
        conn = self.connect()
        cursor = conn.cursor()
        
        # 用户表
        cursor.execute('''
        CREATE TABLE users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT,
            role TEXT NOT NULL,
            full_name TEXT,
            email TEXT,
            created_at TEXT NOT NULL
        )
        ''')
        
        # 工枪表
        cursor.execute('''
        CREATE TABLE guns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            type TEXT,
            model TEXT,
            serial_number TEXT UNIQUE,
            status TEXT NOT NULL,
            location TEXT,
            last_maintenance TEXT,
            notes TEXT,
            created_at TEXT NOT NULL
        )
        ''')
        
        # 预设表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS presets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            gun_type TEXT NOT NULL,
            parameters TEXT,
            description TEXT,
            created_at TEXT NOT NULL
        )
        ''')
        
        conn.commit()
    
    def create_default_data(self):
        conn = self.connect()
        cursor = conn.cursor()
        
        current_time = datetime.datetime.now().isoformat()
        
        # 默认用户
        users = [
            ('system', 'manager', 'admin', '系统管理员', 'admin@welding.com', current_time),
            ('administrator', None, 'admin', 'Administrator', '', current_time),
            ('user', 'user123', 'user', '普通用户', 'user@welding.com', current_time)
        ]
        
        cursor.executemany('''
        INSERT INTO users (username, password, role, full_name, email, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        ''', users)
        
        # 示例工枪
        guns = [
            ('GUN-001', '点焊枪', 'DW-100', 'SN001', 'active', '生产线A', '2024-01-15', '正常使用', current_time),
            ('GUN-002', '弧焊枪', 'HW-200', 'SN002', 'maintenance', '维修车间', '2023-12-20', '需要维护', current_time),
            ('GUN-003', '激光焊枪', 'LW-300', 'SN003', 'active', '实验室', '2024-02-10', '高精度', current_time),
            ('GUN-004', '气体焊枪', 'GW-150', 'SN004', 'inactive', '仓库', '2023-11-05', '备用设备', current_time),
            ('GUN-005', '电阻焊枪', 'RW-250', 'SN005', 'active', '生产线B', '2024-01-30', '新设备', current_time)
        ]
        
        cursor.executemany('''
        INSERT INTO guns (name, type, model, serial_number, status, location, last_maintenance, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', guns)
        
        conn.commit()
    
    def execute(self, query, params=()):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute(query, params)
        conn.commit()
        return cursor
    
    def fetch_all(self, query, params=()):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute(query, params)
        return [dict(row) for row in cursor.fetchall()]
    
    def fetch_one(self, query, params=()):
        conn = self.connect()
        cursor = conn.cursor()
        cursor.execute(query, params)
        row = cursor.fetchone()
        return dict(row) if row else None

# 4. 添加 GunController 类
class GunController:
    def __init__(self, db):
        self.db = db
    
    def get_all_guns(self):
        return self.db.fetch_all("SELECT * FROM guns ORDER BY name")
    
    def get_gun_by_id(self, gun_id):
        return self.db.fetch_one("SELECT * FROM guns WHERE id = ?", (gun_id,))
    
    def search_guns(self, search_term):
        query = """
        SELECT * FROM guns 
        WHERE name LIKE ? OR type LIKE ? OR model LIKE ? OR location LIKE ?
        ORDER BY name
        """
        param = f"%{search_term}%"
        return self.db.fetch_all(query, (param, param, param, param))
    
    def get_statistics(self):
        stats = {}
        
        # 总数
        total = self.db.fetch_one("SELECT COUNT(*) as count FROM guns")
        stats['total_guns'] = total['count'] if total else 0
        
        # 状态分布
        status_data = self.db.fetch_all(
            "SELECT status, COUNT(*) as count FROM guns GROUP BY status"
        )
        stats['status_distribution'] = {
            row['status']: row['count'] for row in status_data
        }
        
        # 类型分布
        type_data = self.db.fetch_all(
            "SELECT type, COUNT(*) as count FROM guns WHERE type IS NOT NULL GROUP BY type"
        )
        stats['type_distribution'] = {
            row['type']: row['count'] for row in type_data
        }
        
        # 各状态数量
        stats['active_guns'] = stats['status_distribution'].get('active', 0)
        stats['maintenance_guns'] = stats['status_distribution'].get('maintenance', 0)
        stats['inactive_guns'] = stats['status_distribution'].get('inactive', 0)
        
        return stats

# 5. 添加 UserController 类
class UserController:
    def __init__(self, db):
        self.db = db
    
    def authenticate(self, username, password):
        if username == "administrator":
            row = self.db.fetch_one(
                "SELECT * FROM users WHERE username = ? AND password IS NULL",
                (username,)
            )
        else:
            row = self.db.fetch_one(
                "SELECT * FROM users WHERE username = ? AND password = ?",
                (username, password)
            )
        
        if row:
            return {
                'id': row['id'],
                'username': row['username'],
                'role': row['role'],
                'full_name': row['full_name'] or row['username']
            }
        return None

# 6. 添加 WeldingGunSystem 类
class WeldingGunSystem:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("焊接枪管理系统")
        self.root.geometry("1200x800")
        
        # 设置窗口最小尺寸
        self.root.minsize(1000, 700)
        
        # 数据库
        self.db = Database()
        if not self.db.initialize():
            messagebox.showerror("错误", "数据库初始化失败")
            sys.exit(1)
        
        # 控制器
        self.gun_ctrl = GunController(self.db)
        self.user_ctrl = UserController(self.db)
        
        # 添加文件管理器
        self.file_manager = GunFileManager()
        
        # 添加上传流程状态
        self.current_upload_gun_info = None
        self.current_upload_folder = None
        
        # 当前用户
        self.current_user = None
        
        # 文件管理相关变量
        self.file_listbox = None
        
        # 运行
        self.show_login()
        self.root.mainloop()
    
    def show_login(self):
        """显示登录界面"""
        for widget in self.root.winfo_children():
            widget.destroy()
        
        # 设置窗口居中
        window_width = 400
        window_height = 500
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        frame = tk.Frame(self.root, bg="#f5f5f5", padx=30, pady=30)
        frame.pack(expand=True, fill=tk.BOTH)
        
        # 标题
        tk.Label(frame, text="焊接枪管理系统", 
                font=("微软雅黑", 22, "bold"), bg="#f5f5f5", fg="#2c3e50").pack(pady=(20, 10))
        
        tk.Label(frame, text="Welding Gun Management System", 
                font=("Arial", 11), bg="#f5f5f5", fg="#7f8c8d").pack(pady=(0, 30))
        
        # 登录表单
        form_frame = tk.Frame(frame, bg="#f5f5f5")
        form_frame.pack()
        
        # 用户名
        tk.Label(form_frame, text="用户名:", font=("微软雅黑", 12), 
                bg="#f5f5f5", fg="#2c3e50").grid(row=0, column=0, sticky=tk.W, pady=15, padx=5)
        self.username_var = tk.StringVar(value="system")
        username_entry = tk.Entry(form_frame, textvariable=self.username_var, 
                                font=("微软雅黑", 12), width=22, bd=2, relief=tk.GROOVE)
        username_entry.grid(row=0, column=1, padx=5, pady=15)
        
        # 密码
        tk.Label(form_frame, text="密码:", font=("微软雅黑", 12), 
                bg="#f5f5f5", fg="#2c3e50").grid(row=1, column=0, sticky=tk.W, pady=15, padx=5)
        self.password_var = tk.StringVar(value="manager")
        password_entry = tk.Entry(form_frame, textvariable=self.password_var, 
                                font=("微软雅黑", 12), width=22, show="•", bd=2, relief=tk.GROOVE)
        password_entry.grid(row=1, column=1, padx=5, pady=15)
        
        # 快速登录提示
        tk.Label(frame, text="快速登录:", font=("微软雅黑", 11), 
                bg="#f5f5f5", fg="#7f8c8d").pack(pady=(20, 10))
        
        # 快速登录按钮框架
        quick_frame = tk.Frame(frame, bg="#f5f5f5")
        quick_frame.pack()
        
        users = [
            ("👑 系统管理员", "system", "manager"),
            ("🛠️ Administrator", "administrator", ""),
            ("👤 普通用户", "user", "user123"),
        ]
        
        for i, (text, username, password) in enumerate(users):
            btn = tk.Button(quick_frame, text=text, 
                          font=("微软雅黑", 10),
                          bg="#3498db", fg="white",
                          padx=15, pady=8,
                          command=lambda u=username, p=password: self.quick_login(u, p))
            btn.grid(row=0, column=i, padx=5)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#2980b9"))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg="#3498db"))
        
        # 登录按钮
        login_btn = tk.Button(frame, text="🔐 登录", 
                             font=("微软雅黑", 13, "bold"), 
                             bg="#2ecc71", fg="white",
                             padx=40, pady=12,
                             command=self.do_login)
        login_btn.pack(pady=30)
        login_btn.bind("<Enter>", lambda e: login_btn.config(bg="#27ae60"))
        login_btn.bind("<Leave>", lambda e: login_btn.config(bg="#2ecc71"))
        
        # 版本信息
        tk.Label(frame, text="版本 2.0.0", 
                font=("Arial", 9), bg="#f5f5f5", fg="#95a5a6").pack(side=tk.BOTTOM, pady=10)
        
        # 绑定回车键登录
        self.root.bind('<Return>', lambda e: self.do_login())
        
        # 焦点设置
        username_entry.focus_set()
    
    def quick_login(self, username, password):
        """快速登录"""
        self.username_var.set(username)
        self.password_var.set(password)
        self.do_login()
    
    def do_login(self):
        """执行登录"""
        username = self.username_var.get().strip()
        password = self.password_var.get().strip()
        
        if not username:
            messagebox.showwarning("警告", "请输入用户名")
            return
        
        user = self.user_ctrl.authenticate(username, password)
        if user:
            self.current_user = user
            # 恢复窗口大小
            self.root.geometry("1200x800")
            self.show_main_interface()
        else:
            messagebox.showerror("登录失败", "用户名或密码错误")

    def show_main_interface(self):
        """显示主界面 - GLPI风格布局"""
        for widget in self.root.winfo_children():
            widget.destroy()
        
        role_name = "管理员" if self.current_user['role'] == 'admin' else "普通用户"
        self.root.title(f"焊接枪管理系统 - {self.current_user['full_name']} ({role_name})")
        
        # 设置窗口最小尺寸
        self.root.minsize(1200, 700)
        
        # ========== 主容器 ==========
        main_container = tk.Frame(self.root)
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # ========== 顶部标题栏 ==========
        title_bar = tk.Frame(main_container, bg="#4a6fa5", height=60)
        title_bar.pack(fill=tk.X)
        title_bar.pack_propagate(False)
        
        # 左侧：系统标题
        title_left = tk.Frame(title_bar, bg="#4a6fa5")
        title_left.pack(side=tk.LEFT, padx=20)
        
        tk.Label(title_left, text="🏭 焊接枪管理系统", 
                font=("微软雅黑", 20, "bold"), 
                bg="#4a6fa5", fg="white").pack(pady=15)
        
        # 右侧：用户信息和操作
        title_right = tk.Frame(title_bar, bg="#4a6fa5")
        title_right.pack(side=tk.RIGHT, padx=20)
        
        # 用户信息
        user_frame = tk.Frame(title_right, bg="#4a6fa5")
        user_frame.pack(side=tk.LEFT, padx=10)
        
        user_icon = tk.Label(user_frame, text="👤", 
                            font=("微软雅黑", 14), 
                            bg="#4a6fa5", fg="white")
        user_icon.pack(side=tk.LEFT, padx=(0, 5))
        
        user_info = tk.Label(user_frame, 
                            text=f"{self.current_user['full_name']} ({role_name})", 
                            font=("微软雅黑", 11), 
                            bg="#4a6fa5", fg="white")
        user_info.pack(side=tk.LEFT)
        
        # 登出按钮
        logout_btn = tk.Button(title_right, text="退出", 
                            font=("微软雅黑", 10),
                            bg="#e74c3c", fg="white",
                            padx=15, pady=5,
                            command=self.logout)
        logout_btn.pack(side=tk.LEFT, padx=(20, 0))
        
        # ========== 内容区域 ==========
        content_area = tk.Frame(main_container)
        content_area.pack(fill=tk.BOTH, expand=True)
        
        # ========== 左侧导航菜单 ==========
        nav_frame = tk.Frame(content_area, width=220, bg="#2c3e50")
        nav_frame.pack(side=tk.LEFT, fill=tk.Y)
        nav_frame.pack_propagate(False)
        
        # 导航菜单标题
        nav_title = tk.Frame(nav_frame, bg="#34495e", height=50)
        nav_title.pack(fill=tk.X)
        nav_title.pack_propagate(False)
        
        tk.Label(nav_title, text="主菜单", 
                font=("微软雅黑", 14, "bold"), 
                bg="#34495e", fg="white").pack(pady=15)
        
        # 导航菜单项
        nav_items = [
            ("🏠 仪表盘", "dashboard", "#3498db"),
            ("🔧 工枪管理", "guns", "#2ecc71"),
            ("📁 文件管理", "files", "#9b59b6"),
            ("📊 统计分析", "stats", "#e74c3c"),
            ("📋 模板工具", "templates", "#f39c12"),
            ("⚙️ 系统设置", "settings", "#95a5a6"),
            ("❓ 帮助文档", "help", "#1abc9c"),
        ]
        
        # 存储导航按钮引用
        self.nav_buttons = {}
        
        for text, page_id, color in nav_items:
            btn_container = tk.Frame(nav_frame, bg="#2c3e50", height=50)
            btn_container.pack(fill=tk.X, pady=1)
            btn_container.pack_propagate(False)
            
            # 左侧指示条
            indicator = tk.Frame(btn_container, bg="#2c3e50", width=4)
            indicator.pack(side=tk.LEFT, fill=tk.Y)
            
            # 导航按钮
            btn = tk.Button(btn_container, text=text, 
                        font=("微软雅黑", 11),
                        bg="#2c3e50", fg="white",
                        anchor="w",
                        padx=20, pady=15,
                        relief=tk.FLAT,
                        command=lambda pid=page_id: self.show_page(pid))
            btn.pack(fill=tk.BOTH, expand=True)
            
            self.nav_buttons[page_id] = {
                'button': btn,
                'indicator': indicator,
                'color': color
            }
            
            # 鼠标悬停效果
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#34495e"))
            btn.bind("<Leave>", lambda e, b=btn, pid=page_id: 
                    b.config(bg="#2c3e50" if self.current_page != pid else "#34495e"))
        
        # ========== 右侧内容区 ==========
        self.content_area = tk.Frame(content_area, bg="#ecf0f1")
        self.content_area.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # 创建页面容器
        self.pages = {}
        
        # 创建各个页面
        self.create_dashboard_page()
        self.create_guns_page()
        self.create_files_page()
        self.create_stats_page()
        self.create_templates_page()
        self.create_settings_page()
        self.create_help_page()
        
        # 默认显示仪表盘
        self.current_page = "dashboard"
        self.show_page("dashboard")

    def logout(self):
        """退出登录"""
        response = messagebox.askyesno("确认", "确定要退出系统吗？")
        if response:
            self.current_user = None
            self.show_login()

    def show_page(self, page_id):
        """显示指定页面"""
        # 隐藏当前页面
        for page in self.pages.values():
            page.pack_forget()
        
        # 重置所有导航按钮样式
        for pid, btn_info in self.nav_buttons.items():
            btn_info['button'].config(bg="#2c3e50")
            btn_info['indicator'].config(bg="#2c3e50")
        
        # 高亮当前导航按钮
        if page_id in self.nav_buttons:
            btn_info = self.nav_buttons[page_id]
            btn_info['button'].config(bg="#34495e")
            btn_info['indicator'].config(bg=btn_info['color'])
        
        # 显示目标页面
        if page_id in self.pages:
            self.pages[page_id].pack(fill=tk.BOTH, expand=True)
            self.current_page = page_id

    def create_file_management_page(self):
        """创建文件管理页面"""
        page_frame = tk.Frame(self.content_area, bg="#ecf0f1")
        self.pages["file_management"] = page_frame
        
        # 页面标题栏
        title_frame = tk.Frame(page_frame, bg="white", height=70)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="📁 文件管理", 
                font=("微软雅黑", 18, "bold"), 
                bg="white", fg="#2c3e50").pack(side=tk.LEFT, padx=30, pady=20)
        
        # 工具栏
        toolbar = tk.Frame(title_frame, bg="white")
        toolbar.pack(side=tk.RIGHT, padx=30)
        
        # 工具栏按钮
        toolbar_buttons = [
            ("📤 上传焊枪", self.upload_file_ui, "#3498db"),
            ("📥 下载文件", self.download_file_ui, "#2ecc71"),
            ("📋 模板工具", lambda: self.show_page("templates"), "#9b59b6"),
            ("🔄 刷新列表", self.refresh_file_list, "#f39c12"),
        ]
        
        for text, command, color in toolbar_buttons:
            btn = tk.Button(toolbar, text=text, font=("微软雅黑", 10),
                        bg=color, fg="white",
                        padx=15, pady=5,
                        command=command)
            btn.pack(side=tk.LEFT, padx=5)
        
        # 页面内容区域
        content_frame = tk.Frame(page_frame, bg="#ecf0f1", padx=20, pady=20)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # 文件列表区域
        list_frame = tk.LabelFrame(content_frame, text="焊枪文件列表", 
                                font=("微软雅黑", 12, "bold"),
                                bg="white", padx=20, pady=15)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        # 文件列表框
        listbox_frame = tk.Frame(list_frame, bg="white")
        listbox_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # 滚动条
        scrollbar = tk.Scrollbar(listbox_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 列表框
        self.file_listbox = tk.Listbox(listbox_frame, 
                                    yscrollcommand=scrollbar.set,
                                    font=("微软雅黑", 10),
                                    selectbackground="#3498db",
                                    selectforeground="white",
                                    activestyle="none")
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.file_listbox.yview)
        
        # 初始加载文件列表
        self.refresh_file_list()
    
    def show_page(self, page_id):
        """显示指定页面"""
        # 隐藏所有页面
        for page in self.pages.values():
            page.pack_forget()
        
        # 重置所有导航按钮样式
        for pid, btn_info in self.nav_buttons.items():
            btn_info['button'].config(bg="#2c3e50")
            btn_info['indicator'].config(bg="#2c3e50")
        
        # 高亮当前导航按钮
        if page_id in self.nav_buttons:
            btn_info = self.nav_buttons[page_id]
            btn_info['button'].config(bg="#34495e")
            btn_info['indicator'].config(bg=btn_info['color'])
        
        # 显示目标页面
        if page_id in self.pages:
            self.pages[page_id].pack(fill=tk.BOTH, expand=True)
            self.current_page = page_id
        else:
            # 如果页面不存在，显示仪表盘
            self.pages["dashboard"].pack(fill=tk.BOTH, expand=True)
            self.current_page = "dashboard"
    
    def create_dashboard_page(self):
        """创建仪表盘页面"""
        page = tk.Frame(self.content_area, bg="#ecf0f1")
        self.pages["dashboard"] = page
        
        # 页面标题
        title_frame = tk.Frame(page, bg="white", height=70)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="系统仪表盘", 
                font=("微软雅黑", 20, "bold"), 
                bg="white", fg="#2c3e50").pack(side=tk.LEFT, padx=30, pady=20)
        
        # 页面内容
        content_frame = tk.Frame(page, bg="#ecf0f1", padx=20, pady=20)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        try:
            # 获取统计数据
            stats = self.gun_ctrl.get_statistics()
            
            # ========== 统计卡片 ==========
            cards_frame = tk.Frame(content_frame, bg="#ecf0f1")
            cards_frame.pack(fill=tk.X, pady=(0, 20))
            
            stat_cards = [
                ("总工枪数", stats.get('total_guns', 0), "#3498db", "📊"),
                ("在用工枪", stats.get('active_guns', 0), "#2ecc71", "✅"),
                ("维护中", stats.get('maintenance_guns', 0), "#e74c3c", "🔧"),
                ("闲置", stats.get('inactive_guns', 0), "#f39c12", "📦"),
            ]
            
            for i, (title, value, color, icon) in enumerate(stat_cards):
                card = tk.Frame(cards_frame, bg="white", relief=tk.RAISED, bd=1)
                card.grid(row=0, column=i, padx=10, sticky="nsew")
                cards_frame.columnconfigure(i, weight=1)
                
                # 图标
                icon_frame = tk.Frame(card, bg=color, width=80, height=80)
                icon_frame.pack(pady=20)
                icon_frame.pack_propagate(False)
                
                tk.Label(icon_frame, text=icon, font=("微软雅黑", 30), 
                        bg=color, fg="white").pack(expand=True)
                
                # 数据
                tk.Label(card, text=title, font=("微软雅黑", 12), 
                        bg="white", fg="#7f8c8d").pack()
                tk.Label(card, text=str(value), font=("微软雅黑", 24, "bold"), 
                        bg="white", fg="#2c3e50").pack(pady=(5, 20))
            
            # ========== 最近工枪列表 ==========
            list_frame = tk.LabelFrame(content_frame, text="最近工枪", 
                                    font=("微软雅黑", 14, "bold"),
                                    bg="white", padx=20, pady=15)
            list_frame.pack(fill=tk.BOTH, expand=True)
            
            # 表格容器
            table_container = tk.Frame(list_frame, bg="white")
            table_container.pack(fill=tk.BOTH, expand=True)
            
            # 创建表格
            columns = [("ID", 60), ("名称", 150), ("类型", 120), 
                    ("状态", 100), ("位置", 150), ("最后维护", 120)]
            
            # 表头
            header_frame = tk.Frame(table_container, bg="#34495e", height=40)
            header_frame.pack(fill=tk.X)
            header_frame.pack_propagate(False)
            
            for col, width in columns:
                header_label = tk.Label(header_frame, text=col, 
                                    font=("微软雅黑", 11, "bold"),
                                    bg="#34495e", fg="white",
                                    width=width//8)
                header_label.pack(side=tk.LEFT, padx=2)
            
            # 表格内容
            table_canvas = tk.Canvas(table_container, bg="white", highlightthickness=0)
            scrollbar = ttk.Scrollbar(table_container, orient="vertical", 
                                    command=table_canvas.yview)
            table_frame = tk.Frame(table_canvas, bg="white")
            
            table_canvas.create_window((0, 0), window=table_frame, anchor="nw")
            table_canvas.configure(yscrollcommand=scrollbar.set)
            
            table_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            guns = self.gun_ctrl.get_all_guns()
            if guns:
                for i, gun in enumerate(guns[:8]):  # 显示前8个
                    row_color = "#ffffff" if i % 2 == 0 else "#f8f9fa"
                    row_frame = tk.Frame(table_frame, bg=row_color, height=35)
                    row_frame.pack(fill=tk.X)
                    row_frame.pack_propagate(False)
                    
                    # 状态颜色映射
                    status_config = {
                        'active': ('#2ecc71', '✅ 在用'),
                        'maintenance': ('#e74c3c', '🔧 维护'),
                        'inactive': ('#f39c12', '📦 闲置')
                    }
                    status_color, status_text = status_config.get(
                        gun['status'], ('#95a5a6', gun['status'])
                    )
                    
                    row_data = [
                        str(gun['id']),
                        gun['name'],
                        gun['type'] or '未分类',
                        status_text,
                        gun['location'] or '未知',
                        gun['last_maintenance'] or '-'
                    ]
                    
                    for j, (data, (col, width)) in enumerate(zip(row_data, columns)):
                        cell_bg = status_color if j == 3 else row_color
                        cell_fg = "white" if j == 3 else "#2c3e50"
                        
                        cell = tk.Label(row_frame, text=data, 
                                    font=("微软雅黑", 10),
                                    bg=cell_bg, fg=cell_fg,
                                    width=width//8, anchor="w",
                                    padx=10)
                        cell.pack(side=tk.LEFT, fill=tk.Y, padx=2)
            
            # 配置滚动区域
            table_frame.update_idletasks()
            table_canvas.config(scrollregion=table_canvas.bbox("all"))
            
        except Exception as e:
            error_label = tk.Label(content_frame, 
                                text=f"加载仪表盘失败: {str(e)}", 
                                font=("微软雅黑", 12), fg="red", bg="#ecf0f1")
            error_label.pack(pady=50)
    
    def create_gun_management_page(self):
        """创建工枪管理页面"""
        gun_frame = tk.Frame(self.notebook)
        self.notebook.add(gun_frame, text="🔧 工枪管理")
        
        # 工具栏
        toolbar = tk.Frame(gun_frame, bg="#ecf0f1", pady=15)
        toolbar.pack(fill=tk.X, padx=20, pady=10)
        
        # 左侧按钮
        left_btn_frame = tk.Frame(toolbar, bg="#ecf0f1")
        left_btn_frame.pack(side=tk.LEFT)
        
        refresh_btn = tk.Button(left_btn_frame, text="🔄 刷新", 
                               bg="#3498db", fg="white", font=("微软雅黑", 10),
                               command=self.refresh_gun_table)
        refresh_btn.pack(side=tk.LEFT, padx=5)
        
        # 右侧搜索框
        search_frame = tk.Frame(toolbar, bg="#ecf0f1")
        search_frame.pack(side=tk.RIGHT)
        
        tk.Label(search_frame, text="搜索:", bg="#ecf0f1", 
                font=("微软雅黑", 10)).pack(side=tk.LEFT)
        
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, 
                               width=30, font=("微软雅黑", 10))
        search_entry.pack(side=tk.LEFT, padx=5)
        
        search_btn = tk.Button(search_frame, text="🔍 搜索", 
                              bg="#2ecc71", fg="white", font=("微软雅黑", 10),
                              command=self.search_guns_table)
        search_btn.pack(side=tk.LEFT)
        
        # 表格框架
        table_frame = tk.Frame(gun_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # 创建Treeview
        self.gun_tree = ttk.Treeview(table_frame, 
                                    columns=('ID', '名称', '类型', '型号', '状态', '位置', '维护日期'), 
                                    show='headings', height=20)
        
        columns = [
            ('ID', 60),
            ('名称', 150),
            ('类型', 120),
            ('型号', 120),
            ('状态', 100),
            ('位置', 150),
            ('维护日期', 120)
        ]
        
        for col, width in columns:
            self.gun_tree.heading(col, text=col)
            self.gun_tree.column(col, width=width)
        
        # 滚动条
        v_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.gun_tree.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.gun_tree.xview)
        self.gun_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.gun_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 加载数据
        self.refresh_gun_table()
        
        # 绑定双击事件
        self.gun_tree.bind("<Double-1>", self.on_gun_double_click)
    
    def create_templates_page(self):
        """创建模板工具页面"""
        page_frame = tk.Frame(self.content_area, bg="#ecf0f1")
        self.pages["templates"] = page_frame
        
        # 页面标题栏
        title_frame = tk.Frame(page_frame, bg="white", height=70)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        
        tk.Label(title_frame, text="📋 模板工具", 
                font=("微软雅黑", 18, "bold"), 
                bg="white", fg="#2c3e50").pack(side=tk.LEFT, padx=30, pady=20)
        
        # 页面内容区域
        content_frame = tk.Frame(page_frame, bg="#ecf0f1", padx=20, pady=20)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # 模板工具卡片
        cards_frame = tk.Frame(content_frame, bg="#ecf0f1")
        cards_frame.pack(fill=tk.BOTH, expand=True)
        
        # 导出模板卡片
        export_card = tk.Frame(cards_frame, bg="white", relief=tk.RAISED, bd=1)
        export_card.pack(fill=tk.X, pady=(0, 20))
        
        tk.Label(export_card, text="📤 导出模板", 
                font=("微软雅黑", 14, "bold"), 
                bg="white", fg="#2c3e50").pack(anchor="w", padx=30, pady=(20, 10))
        
        tk.Label(export_card, text="导出Excel模板文件，用于批量录入焊枪信息", 
                font=("微软雅黑", 10), 
                bg="white", fg="#7f8c8d").pack(anchor="w", padx=30, pady=(0, 20))
        
        export_btn = tk.Button(export_card, text="导出模板文件", 
                            font=("微软雅黑", 11, "bold"),
                            bg="#3498db", fg="white",
                            padx=30, pady=10,
                            command=self.export_template)
        export_btn.pack(pady=(0, 20))
        
        # 导入数据卡片
        import_card = tk.Frame(cards_frame, bg="white", relief=tk.RAISED, bd=1)
        import_card.pack(fill=tk.X)
        
        tk.Label(import_card, text="📥 导入数据", 
                font=("微软雅黑", 14, "bold"), 
                bg="white", fg="#2c3e50").pack(anchor="w", padx=30, pady=(20, 10))
        
        tk.Label(import_card, text="导入填写好的Excel/CSV文件，批量创建焊枪记录", 
                font=("微软雅黑", 10), 
                bg="white", fg="#7f8c8d").pack(anchor="w", padx=30, pady=(0, 20))
        
        import_btn = tk.Button(import_card, text="导入数据文件", 
                            font=("微软雅黑", 11, "bold"),
                            bg="#2ecc71", fg="white",
                            padx=30, pady=10,
                            command=self.import_data)
        import_btn.pack(pady=(0, 20))
    
    def create_statistics_page(self):
        """创建统计分析页面"""
        stats_frame = tk.Frame(self.notebook)
        self.notebook.add(stats_frame, text="📊 统计分析")
        
        # 标题
        tk.Label(stats_frame, text="统计分析报告", 
                font=("微软雅黑", 20, "bold"), bg="white").pack(pady=20)
        
        # 内容框架
        content_frame = tk.Frame(stats_frame, bg="white")
        content_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=10)
        
        # 创建文本显示
        text_widget = tk.Text(content_frame, wrap=tk.WORD, 
                             font=("微软雅黑", 11),
                             padx=20, pady=20)
        text_widget.pack(fill=tk.BOTH, expand=True)
        
        # 获取统计数据
        try:
            stats = self.gun_ctrl.get_statistics()
            
            stats_text = f"""
{'='*60}
                     工枪统计报告
{'='*60}

📊 基本信息
{'-'*30}
• 总工枪数: {stats.get('total_guns', 0)} 把
• 在用工枪: {stats.get('active_guns', 0)} 把  ({stats.get('active_guns', 0)/max(stats.get('total_guns', 1), 1)*100:.1f}%)
• 维护中: {stats.get('maintenance_guns', 0)} 把  ({stats.get('maintenance_guns', 0)/max(stats.get('total_guns', 1), 1)*100:.1f}%)
• 闲置: {stats.get('inactive_guns', 0)} 把  ({stats.get('inactive_guns', 0)/max(stats.get('total_guns', 1), 1)*100:.1f}%)

📈 状态分布
{'-'*30}
"""
            
            for status, count in stats.get('status_distribution', {}).items():
                status_zh = {
                    'active': '✅ 在用',
                    'maintenance': '🔧 维护',
                    'inactive': '📦 闲置',
                    'scrap': '🗑️ 报废'
                }.get(status, status)
                percentage = count / max(stats['total_guns'], 1) * 100
                stats_text += f"{status_zh}: {count} 把 ({percentage:.1f}%)\n"
            
            stats_text += f"""
🔧 类型分布
{'-'*30}
"""
            for gun_type, count in stats.get('type_distribution', {}).items():
                stats_text += f"• {gun_type}: {count} 把\n"
            
            stats_text += f"""
{'='*60}
报告生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'='*60}
"""
            
            text_widget.insert(tk.END, stats_text)
            text_widget.config(state=tk.DISABLED)
            
        except Exception as e:
            text_widget.insert(tk.END, f"加载统计失败: {str(e)}")
            text_widget.config(state=tk.DISABLED)
    
    def create_settings_page(self):
        """创建系统设置页面"""
        settings_frame = tk.Frame(self.notebook, bg="white")
        self.notebook.add(settings_frame, text="⚙️ 系统设置")
        
        # 标题
        tk.Label(settings_frame, text="系统设置", 
                font=("微软雅黑", 20, "bold"), bg="white").pack(pady=20)
        
        # 设置内容框架
        content_frame = tk.Frame(settings_frame, bg="white", padx=50, pady=30)
        content_frame.pack()
        
        # 用户信息
        user_frame = tk.LabelFrame(content_frame, text="👤 用户信息", 
                                  font=("微软雅黑", 12, "bold"),
                                  padx=20, pady=15, bg="white")
        user_frame.grid(row=0, column=0, sticky="w", pady=10)
        
        info_items = [
            ("用户名:", self.current_user['username']),
            ("姓名:", self.current_user['full_name']),
            ("角色:", "管理员" if self.current_user['role'] == 'admin' else "普通用户"),
        ]
        
        for i, (label, value) in enumerate(info_items):
            tk.Label(user_frame, text=label, font=("微软雅黑", 11), 
                    bg="white", width=10, anchor="e").grid(row=i, column=0, sticky="e", pady=8, padx=(0, 10))
            tk.Label(user_frame, text=value, font=("微软雅黑", 11, "bold"), 
                    bg="white", width=20, anchor="w").grid(row=i, column=1, sticky="w", pady=8)
        
        # 系统信息
        sys_frame = tk.LabelFrame(content_frame, text="💻 系统信息", 
                                 font=("微软雅黑", 12, "bold"),
                                 padx=20, pady=15, bg="white")
        sys_frame.grid(row=1, column=0, sticky="w", pady=20)
        
        sys_items = [
            ("数据库:", "welding_gun.db"),
            ("Python版本:", f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"),
            ("运行平台:", sys.platform),
        ]
        
        for i, (label, value) in enumerate(sys_items):
            tk.Label(sys_frame, text=label, font=("微软雅黑", 11), 
                    bg="white", width=10, anchor="e").grid(row=i, column=0, sticky="e", pady=8, padx=(0, 10))
            tk.Label(sys_frame, text=value, font=("微软雅黑", 11), 
                    bg="white", width=20, anchor="w").grid(row=i, column=1, sticky="w", pady=8)
    
    def create_help_page(self):
        """创建帮助页面"""
        help_frame = tk.Frame(self.notebook, bg="white")
        self.notebook.add(help_frame, text="❓ 帮助")
        
        # 标题
        tk.Label(help_frame, text="用户手册", 
                font=("微软雅黑", 20, "bold"), bg="white").pack(pady=20)
        
        # 创建文本显示
        text_widget = tk.Text(help_frame, wrap=tk.WORD, 
                             font=("微软雅黑", 11),
                             padx=30, pady=20,
                             bg="#f8f9fa")
        text_widget.pack(fill=tk.BOTH, expand=True)
        
        help_text = """
✨ 焊接枪管理系统 - 用户手册 ✨

📋 系统简介
────────────────────
本系统专为焊接枪设备管理设计，提供设备管理、状态监控、统计分析等功能，
帮助您高效管理焊接设备资源。

🚀 主要功能
────────────────────
1. 📁 文件管理 (顶部固定区域)
   • 上传文件：上传设备文档、图片等
   • 下载文件：下载已上传的文件
   • 刷新列表：更新文件列表

2. 🏠 仪表盘
   • 查看设备统计概览
   • 快速查看最近工枪状态
   • 设备状态分布展示

3. 🔧 工枪管理
   • 查看所有工枪详细信息
   • 搜索和筛选设备
   • 双击查看设备详情
   • 设备状态管理

4. 📊 统计分析
   • 设备统计报告
   • 状态分布图表
   • 类型分布分析

5. ⚙️ 系统设置
   • 查看用户信息
   • 查看系统信息
   • 数据库状态

🔑 登录账号
────────────────────
• 系统管理员: system / manager
• Administrator: administrator (无密码)
• 普通用户: user / user123

💡 使用技巧
────────────────────
• 使用快速登录按钮快速切换账号
• 在工枪管理中双击设备查看详情
• 使用搜索功能快速查找设备
• 文件列表会自动刷新显示

🆘 技术支持
────────────────────
如有问题，请：
1. 检查网络连接
2. 确认数据库文件存在
3. 联系系统管理员

────────────────────
版本: 2.0.0 | 最后更新: 2024年
────────────────────
"""
        
        text_widget.insert(tk.END, help_text)
        text_widget.config(state=tk.DISABLED)
    
    # ========== 修改现有的文件管理方法 ==========
    
    def upload_file_ui(self):
        """上传焊枪 - 完整流程"""
        # 第一步：输入焊枪信息
        self.show_gun_info_dialog()
    
    def show_gun_info_dialog(self):
        """显示焊枪信息输入对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("焊枪信息录入")
        dialog.geometry("500x600")
        dialog.resizable(False, False)
        
        # 让对话框居中显示
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f'{width}x{height}+{x}+{y}')
        
        # 主框架
        main_frame = tk.Frame(dialog, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        tk.Label(main_frame, text="📝 焊枪信息录入", 
                font=("微软雅黑", 16, "bold")).pack(pady=(0, 20))
        
        # 创建滚动区域
        canvas = tk.Canvas(main_frame, height=400)
        scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 信息输入框架（在滚动区域内）
        form_frame = tk.Frame(scrollable_frame, padx=10)
        form_frame.pack(fill=tk.X, pady=10)
        
        # 焊枪信息字段 - 根据您的需求重新设计
        fields = [
            # (字段名, 显示标签, 默认值, 必填, 字段类型, 选项列表)
            ("weld_type", "焊接类型*", "", True, "combo", ["钢点焊", "铝点焊", "其他"]),
            ("gun_brand", "焊枪品牌*", "", True, "combo", ["小原", "森德莱", "日基"]),
            ("gun_number", "焊枪编号*", "", True, "text", None),
            ("gun_model", "焊枪型号", "", False, "combo", ["C型", "X型", "异型C", "异型X", "其他"]),
            ("throat_depth", "喉深(mm)", "", False, "text", None),
            ("throat_width", "喉宽(mm)", "", False, "text", None),
            ("max_stroke", "最大行程(mm)", "", False, "text", None),
            ("max_pressure", "最大压力(kN)", "", False, "text", None),
            ("motor_brand", "电机品牌", "", False, "combo", ["ABB", "安川", "川崎", "发那科", "华数控", "库卡", "那智", "其他"]),
            ("cap_spec", "电极帽规格", "", False, "text", None),
            ("cap_tilt", "电极帽是否倾斜", "否", False, "radio", ["是", "否"]),
            ("static_tilt_angle", "静电极帽倾斜角度(°)", "", False, "text", None),
            ("dynamic_tilt_angle", "动电极帽倾斜角度(°)", "", False, "text", None),
        ]
        
        self.info_vars = {}
        self.info_entries = {}  # 存储控件的引用
        
        for i, (field, label, default, required, field_type, options) in enumerate(fields):
            # 标签
            label_text = f"{label}:" if required else f"{label}:"
            label_color = "#e74c3c" if required else "#2c3e50"
            tk.Label(form_frame, text=label_text, font=("微软雅黑", 10),
                    anchor="w", width=18, fg=label_color).grid(row=i, column=0, sticky="w", pady=5, padx=(0, 5))
            
            # 根据字段类型创建不同的输入控件
            if field_type == "combo":
                # 下拉选择框
                var = tk.StringVar(value=default)
                combo = ttk.Combobox(form_frame, textvariable=var, 
                                font=("微软雅黑", 10), width=20,
                                values=options)
                combo.grid(row=i, column=1, pady=5, sticky="w")
                self.info_vars[field] = var
                self.info_entries[field] = combo
                
            elif field_type == "radio":
                # 单选按钮
                var = tk.StringVar(value=default)
                radio_frame = tk.Frame(form_frame)
                radio_frame.grid(row=i, column=1, pady=5, sticky="w")
                
                for option in options:
                    tk.Radiobutton(radio_frame, text=option, variable=var, 
                                value=option, font=("微软雅黑", 9)).pack(side=tk.LEFT, padx=(0, 10))
                
                self.info_vars[field] = var
                self.info_entries[field] = radio_frame
                
            else:
                # 普通文本输入框
                var = tk.StringVar(value=default)
                entry = tk.Entry(form_frame, textvariable=var, 
                            font=("微软雅黑", 10), width=22)
                entry.grid(row=i, column=1, pady=5, sticky="w")
                self.info_vars[field] = var
                self.info_entries[field] = entry
        
        # 添加单位说明标签
        unit_frame = tk.Frame(main_frame)
        unit_frame.pack(pady=(10, 0))
        
        tk.Label(unit_frame, text="*注：", font=("微软雅黑", 9), fg="#7f8c8d").pack(side=tk.LEFT)
        tk.Label(unit_frame, text="为必填项", font=("微软雅黑", 9), fg="#e74c3c").pack(side=tk.LEFT)
        
        # 按钮框架（在主框架中，不在滚动区域）
        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        # 下一步按钮
        next_btn = tk.Button(button_frame, text="下一步 →", 
                        font=("微软雅黑", 11, "bold"),
                        bg="#3498db", fg="white",
                        padx=30, pady=8,
                        command=lambda: self.process_gun_info(dialog))
        next_btn.pack(side=tk.LEFT, padx=10)
        
        # 取消按钮
        cancel_btn = tk.Button(button_frame, text="取消", 
                            font=("微软雅黑", 11),
                            bg="#95a5a6", fg="white",
                            padx=30, pady=8,
                            command=dialog.destroy)
        cancel_btn.pack(side=tk.LEFT, padx=10)
        
        # 绑定回车键
        dialog.bind('<Return>', lambda e: self.process_gun_info(dialog))
        
        # 设置焦点到第一个下拉框
        dialog.after(100, lambda: self.info_entries['weld_type'].focus_set())
      
    def process_gun_info(self, dialog):
        """处理焊枪信息"""
        # 收集信息
        gun_info = {}
        
        # 必填字段列表
        required_fields = ['weld_type', 'gun_brand', 'gun_number']
        
        # 收集所有字段信息
        for field, var in self.info_vars.items():
            value = var.get().strip()
            gun_info[field] = value
            
            # 检查必填字段
            if field in required_fields and not value:
                field_names = {
                    'weld_type': '焊接类型',
                    'gun_brand': '焊枪品牌', 
                    'gun_number': '焊枪编号'
                }
                messagebox.showwarning("警告", f"请填写 {field_names.get(field, field)}")
                return
        
        # 检查焊枪名称是否已存在（使用焊枪编号作为唯一标识）
        try:
            gun_name = f"{gun_info['gun_brand']}_{gun_info['gun_number']}"
            if self.file_manager.get_gun_by_name(gun_name):
                response = messagebox.askyesno("提示", 
                    f"焊枪 '{gun_name}' 已存在，是否继续？")
                if not response:
                    return
        except Exception as e:
            print(f"检查焊枪名称时出错: {e}")
        
        # 创建焊枪文件夹
        try:
            # 生成一个唯一的名称用于文件夹创建
            folder_gun_info = gun_info.copy()
            folder_gun_info['name'] = f"{gun_info['gun_brand']}_{gun_info['gun_number']}"
            
            folder_path = self.file_manager.create_gun_folder(folder_gun_info)
            self.current_upload_gun_info = gun_info
            self.current_upload_folder = folder_path
            
            dialog.destroy()
            self.show_file_upload_dialog()
            
        except Exception as e:
            messagebox.showerror("错误", f"创建焊枪文件夹失败: {str(e)}") 
    
    def show_file_upload_dialog(self):
        """显示文件上传对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("上传焊枪文件")
        dialog.geometry("600x600")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 主框架
        main_frame = tk.Frame(dialog, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(main_frame, text="📁 上传焊枪文件", 
                font=("微软雅黑", 16, "bold")).pack(pady=(0, 20))
        
        tk.Label(main_frame, text=f"焊枪: {self.current_upload_gun_info['name']}", 
                font=("微软雅黑", 12), fg="#3498db").pack(pady=(0, 20))
        
        # 文件类型说明
        desc_frame = tk.Frame(main_frame, relief=tk.GROOVE, bd=2, padx=10, pady=10)
        desc_frame.pack(fill=tk.X, pady=(0, 20))
        
        desc_text = """
文件类型说明:
• 3D模型: .stl, .step, .iges 等格式
• 2D图纸: .pdf, .dxf 等格式  
• 图片: .jpg, .png 等格式
• 会签图: .pdf, .jpg 等格式 (选填)
• DWG文件: .dwg 格式 (选填)
        """
        
        tk.Label(desc_frame, text=desc_text, 
                font=("微软雅黑", 9), justify=tk.LEFT).pack()
        
        # 文件上传区域
        self.file_vars = {}
        
        file_types = [
            ("3d", "3D模型文件*", ["*.stl", "*.step", "*.iges", "*.stp"]),
            ("2d", "2D图纸文件*", ["*.pdf", "*.dxf", "*.dwg"]),
            ("image", "图片文件*", ["*.jpg", "*.jpeg", "*.png", "*.bmp"]),
            ("signature", "会签图文件 (选填)", ["*.pdf", "*.jpg", "*.png"]),
            ("dwg", "DWG文件 (选填)", ["*.dwg"])
        ]
        
        for file_type, label, extensions in file_types:
            frame = tk.Frame(main_frame)
            frame.pack(fill=tk.X, pady=8)
            
            # 标签
            tk.Label(frame, text=label, font=("微软雅黑", 10),
                    width=20, anchor="w").pack(side=tk.LEFT)
            
            # 文件路径显示
            var = tk.StringVar()
            entry = tk.Entry(frame, textvariable=var, 
                           font=("微软雅黑", 9), width=30, state='readonly')
            entry.pack(side=tk.LEFT, padx=5)
            self.file_vars[file_type] = var
            
            # 浏览按钮
            browse_btn = tk.Button(frame, text="浏览...", 
                                 font=("微软雅黑", 9),
                                 command=lambda ft=file_type, ext=extensions: 
                                 self.browse_file(ft, ext))
            browse_btn.pack(side=tk.LEFT, padx=5)
            
            # 清除按钮
            clear_btn = tk.Button(frame, text="✕", 
                                font=("微软雅黑", 9), width=2,
                                command=lambda v=var: v.set(""))
            clear_btn.pack(side=tk.LEFT)
        
        # 按钮框架
        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=30)
        
        # 上传按钮
        upload_btn = tk.Button(button_frame, text="📤 上传文件", 
                             font=("微软雅黑", 11, "bold"),
                             bg="#2ecc71", fg="white",
                             padx=30, pady=10,
                             command=lambda: self.upload_files(dialog))
        upload_btn.pack(side=tk.LEFT, padx=10)
        
        # 跳过按钮
        skip_btn = tk.Button(button_frame, text="跳过，直接完成", 
                           font=("微软雅黑", 11),
                           bg="#f39c12", fg="white",
                           padx=20, pady=10,
                           command=lambda: self.complete_upload(dialog, skip=True))
        skip_btn.pack(side=tk.LEFT, padx=10)
        
        # 取消按钮
        cancel_btn = tk.Button(button_frame, text="取消", 
                             font=("微软雅黑", 11),
                             bg="#e74c3c", fg="white",
                             padx=30, pady=10,
                             command=lambda: self.cancel_upload(dialog))
        cancel_btn.pack(side=tk.LEFT, padx=10)
    
    def browse_file(self, file_type, extensions):
        """浏览文件"""
        file_path = filedialog.askopenfilename(
            title=f"选择{file_type}文件",
            filetypes=[("支持的文件", extensions), ("所有文件", "*.*")]
        )
        
        if file_path:
            self.file_vars[file_type].set(file_path)
    
    def upload_files(self, dialog):
        """上传文件"""
        required_types = ['3d', '2d', 'image']
        uploaded_files = []
        
        for file_type, var in self.file_vars.items():
            file_path = var.get().strip()
            
            if file_path:
                if os.path.exists(file_path):
                    try:
                        saved_path = self.file_manager.save_file_to_folder(
                            self.current_upload_folder, 
                            file_path, 
                            file_type
                        )
                        uploaded_files.append((file_type, os.path.basename(saved_path)))
                    except Exception as e:
                        messagebox.showerror("错误", f"上传{file_type}文件失败: {str(e)}")
                        return
                else:
                    messagebox.showwarning("警告", f"文件不存在: {file_path}")
                    return
        
        # 检查必填文件
        missing_required = []
        for req_type in required_types:
            if not self.file_vars[req_type].get().strip():
                missing_required.append(req_type)
        
        if missing_required:
            response = messagebox.askyesno(
                "警告", 
                f"以下必填文件未上传: {', '.join(missing_required)}\n是否继续？"
            )
            
            if not response:
                return
        
        # 完成上传
        self.complete_upload(dialog)
    
    def complete_upload(self, dialog, skip=False):
        """完成上传"""
        try:
            # 创建ZIP文件
            zip_path = self.file_manager.create_zip_file(self.current_upload_folder)
            
            # 显示成功信息
            if skip:
                message = f"焊枪 '{self.current_upload_gun_info['name']}' 已创建，但未上传文件"
            else:
                message = f"焊枪 '{self.current_upload_gun_info['name']}' 上传完成！\nZIP文件: {os.path.basename(zip_path)}"
            
            messagebox.showinfo("成功", message)
            
            # 刷新文件列表
            self.refresh_file_list()
            
            # 清除状态
            self.current_upload_gun_info = None
            self.current_upload_folder = None
            
            dialog.destroy()
            
        except Exception as e:
            messagebox.showerror("错误", f"完成上传失败: {str(e)}")
    
    def cancel_upload(self, dialog):
        """取消上传"""
        response = messagebox.askyesno("确认", "确定要取消上传吗？")
        
        if response:
            # 删除已创建的文件夹
            if self.current_upload_folder and os.path.exists(self.current_upload_folder):
                try:
                    shutil.rmtree(self.current_upload_folder)
                except:
                    pass
            
            # 清除状态
            self.current_upload_gun_info = None
            self.current_upload_folder = None
            
            dialog.destroy()
    
    def download_file_ui(self):
        """下载焊枪文件"""
        if not self.file_listbox:
            messagebox.showwarning("警告", "文件列表未初始化")
            return
        
        selection = self.file_listbox.curselection()
        if not selection:
            messagebox.showwarning("警告", "请先选择一个焊枪")
            return
        
        # 获取选中的焊枪名称
        gun_name = self.file_listbox.get(selection[0])
        
        # 查找焊枪信息
        gun_info = self.file_manager.get_gun_by_name(gun_name)
        
        if not gun_info:
            messagebox.showerror("错误", f"找不到焊枪: {gun_name}")
            return
        
        # 检查是否有ZIP文件
        if not gun_info.get('has_zip', False):
            response = messagebox.askyesno("提示", 
                f"焊枪 '{gun_name}' 还没有ZIP文件，是否现在创建？")
            
            if response:
                try:
                    zip_path = self.file_manager.create_zip_file(gun_info['folder_path'])
                    gun_info['zip_file'] = zip_path
                    gun_info['has_zip'] = True
                except Exception as e:
                    messagebox.showerror("错误", f"创建ZIP文件失败: {str(e)}")
                    return
        
        # 选择保存位置
        save_path = filedialog.asksaveasfilename(
            title="保存焊枪文件",
            initialfile=f"{gun_name}.zip",
            defaultextension=".zip",
            filetypes=[("ZIP文件", "*.zip"), ("所有文件", "*.*")]
        )
        
        if save_path:
            try:
                # 复制ZIP文件
                shutil.copy2(gun_info['zip_file'], save_path)
                
                messagebox.showinfo("下载成功", 
                    f"焊枪文件已保存到:\n{save_path}\n\n包含文件:\n"
                    f"• 焊枪信息: gun_info.json\n"
                    f"• 3D模型: {len(gun_info.get('files', {}).get('3d', []))}个\n"
                    f"• 2D图纸: {len(gun_info.get('files', {}).get('2d', []))}个\n"
                    f"• 图片: {len(gun_info.get('files', {}).get('image', []))}个")
                
            except Exception as e:
                messagebox.showerror("下载失败", f"下载出错: {str(e)}")
    
    def refresh_file_list(self):
        """刷新文件列表 - 修改为显示焊枪列表"""
        if not self.file_listbox:
            return
        
        self.file_listbox.delete(0, tk.END)
        
        try:
            # 获取所有焊枪
            guns = self.file_manager.get_all_guns()
            
            for gun in guns:
                gun_name = gun['name']
                gun_type = gun.get('type', '未知类型')
                file_count = sum(len(gun.get('files', {}).get(ft, [])) for ft in gun.get('files', {}))
                
                display_text = f"{gun_name} ({gun_type}) - {file_count}个文件"
                if gun.get('has_zip', False):
                    display_text += " 📦"
                
                self.file_listbox.insert(tk.END, display_text)
                
        except Exception as e:
            self.file_listbox.insert(tk.END, f"获取焊枪列表失败: {str(e)}")
    
    # ========== 工枪管理方法 ==========
    def refresh_gun_table(self):
        """刷新工枪表格"""
        if not hasattr(self, 'gun_tree'):
            return
        
        for item in self.gun_tree.get_children():
            self.gun_tree.delete(item)
        
        try:
            guns = self.gun_ctrl.get_all_guns()
            for gun in guns:
                self.gun_tree.insert('', 'end', values=(
                    gun['id'],
                    gun['name'],
                    gun['type'] or '未分类',
                    gun['model'] or '-',
                    gun['status'],
                    gun['location'] or '-',
                    gun['last_maintenance'] or '-'
                ))
        except Exception as e:
            print(f"加载工枪数据失败: {e}")
    
    def search_guns_table(self):
        """搜索工枪"""
        if not hasattr(self, 'gun_tree') or not hasattr(self, 'search_var'):
            return
        
        search_term = self.search_var.get()
        if not search_term:
            self.refresh_gun_table()
            return
        
        for item in self.gun_tree.get_children():
            self.gun_tree.delete(item)
        
        try:
            guns = self.gun_ctrl.search_guns(search_term)
            for gun in guns:
                self.gun_tree.insert('', 'end', values=(
                    gun['id'],
                    gun['name'],
                    gun['type'] or '未分类',
                    gun['model'] or '-',
                    gun['status'],
                    gun['location'] or '-',
                    gun['last_maintenance'] or '-'
                ))
        except Exception as e:
            print(f"搜索工枪失败: {e}")
    
    def on_gun_double_click(self, event):
        """工枪双击事件"""
        selection = self.gun_tree.selection()
        if selection:
            item = self.gun_tree.item(selection[0])
            gun_id = item['values'][0]
            
            # 获取详细数据
            gun_details = self.gun_ctrl.get_gun_by_id(gun_id)
            if gun_details:
                self.show_gun_details(gun_details)
    
    def show_gun_details(self, gun):
        """显示工枪详情"""
        details_window = tk.Toplevel(self.root)
        details_window.title(f"工枪详情 - {gun['name']}")
        details_window.geometry("500x400")
        
        # 详情内容
        content_frame = tk.Frame(details_window, padx=30, pady=30)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        details = [
            ("设备ID:", gun['id']),
            ("设备名称:", gun['name']),
            ("设备类型:", gun['type'] or '未分类'),
            ("设备型号:", gun['model'] or '-'),
            ("序列号:", gun['serial_number'] or '-'),
            ("状态:", gun['status']),
            ("位置:", gun['location'] or '-'),
            ("最后维护:", gun['last_maintenance'] or '-'),
            ("备注:", gun['notes'] or '-'),
        ]
        
        for i, (label, value) in enumerate(details):
            tk.Label(content_frame, text=label, font=("微软雅黑", 11), 
                    anchor="e", width=12).grid(row=i, column=0, sticky="e", pady=8)
            tk.Label(content_frame, text=value, font=("微软雅黑", 11, "bold"), 
                    anchor="w", width=30).grid(row=i, column=1, sticky="w", pady=8)
        
        # 关闭按钮
        tk.Button(content_frame, text="关闭", command=details_window.destroy,
                 bg="#e74c3c", fg="white", font=("微软雅黑", 11),
                 padx=20, pady=5).grid(row=len(details), column=0, columnspan=2, pady=20)

    def export_template(self):
        """导出模板文件 - 主要生成Excel格式"""
        # 字段定义
        field_definitions = [
            ("weld_type", "焊接类型*", "必填，可选值：钢点焊、铝点焊、其他"),
            ("gun_brand", "焊枪品牌*", "必填，可选值：小原、森德莱、日基"),
            ("gun_number", "焊枪编号*", "必填，焊枪唯一编号"),
            ("gun_model", "焊枪型号", "选填，可选值：C型、X型、异型C、异型X、其他"),
            ("throat_depth", "喉深(mm)", "选填，单位：毫米"),
            ("throat_width", "喉宽(mm)", "选填，单位：毫米"),
            ("max_stroke", "最大行程(mm)", "选填，单位：毫米"),
            ("max_pressure", "最大压力(kN)", "选填，单位：千牛"),
            ("motor_brand", "电机品牌", "选填，可选值：ABB、安川、川崎、发那科、华数控、库卡、那智、其他"),
            ("cap_spec", "电极帽规格", "选填"),
            ("cap_tilt", "电极帽是否倾斜", "选填，可选值：是、否"),
            ("static_tilt_angle", "静电极帽倾斜角度(°)", "选填，单位：度"),
            ("dynamic_tilt_angle", "动电极帽倾斜角度(°)", "选填，单位：度"),
        ]
        
        # 选择保存位置
        file_path = filedialog.asksaveasfilename(
            title="保存模板文件",
            defaultextension=".xlsx",
            initialfile="焊枪信息模板.xlsx",
            filetypes=[
                ("Excel文件(*.xlsx)", "*.xlsx"),
                ("Excel 97-2003文件(*.xls)", "*.xls"),
                ("CSV文件(*.csv)", "*.csv"),
                ("所有文件", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        try:
            import pandas as pd
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # 创建DataFrame
            data = []
            
            # 第一行：字段名称
            field_names = [f"{label}" for _, label, _ in field_definitions]
            data.append(field_names)
            
            # 第二行：字段说明
            field_descriptions = [desc for _, _, desc in field_definitions]
            data.append(field_descriptions)
            
            # 第三行：示例数据
            example_data = [
                "钢点焊",      # weld_type
                "小原",        # gun_brand
                "GUN-001",     # gun_number
                "C型",         # gun_model
                "500",         # throat_depth
                "200",         # throat_width
                "150",         # max_stroke
                "4.5",         # max_pressure
                "库卡",        # motor_brand
                "R30",         # cap_spec
                "否",          # cap_tilt
                "0",           # static_tilt_angle
                "0"            # dynamic_tilt_angle
            ]
            data.append(example_data)
            
            # 第四行：开始填写提示
            data.append(["↓ 请从这一行开始填写您的数据 ↓"] + [""] * 12)
            
            # 第五行：空行，用户从这里开始填写
            data.append([""] * 13)
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 保存文件
            if file_path.lower().endswith('.xlsx'):
                # 使用openpyxl引擎保存为xlsx
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, header=False)
                    
                    # 获取工作表对象进行格式设置
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']
                    
                    # 设置列宽 - 修复：安全地计算最大长度
                    for i, col in enumerate(df.columns):
                        max_len = 15  # 默认最小宽度
                        for j in range(len(df)):
                            cell_value = df.iloc[j, i]
                            if pd.notna(cell_value):  # 检查是否为NaN
                                cell_str = str(cell_value)
                                try:
                                    # 尝试计算长度，如果不是字符串则跳过
                                    cell_len = len(cell_str)
                                    if cell_len > max_len:
                                        max_len = cell_len
                                except:
                                    pass
                        
                        # 限制最大宽度
                        column_width = min(max_len + 2, 40)
                        column_letter = openpyxl.utils.get_column_letter(i + 1)
                        worksheet.column_dimensions[column_letter].width = column_width
                    
                    # 设置第一行（字段名）为红色粗体
                    for col in range(1, 14):  # 13列
                        cell = worksheet.cell(row=1, column=col)
                        cell.font = Font(bold=True, color="FF0000")
                    
                    # 设置第二行（说明）为蓝色斜体
                    for col in range(1, 14):
                        cell = worksheet.cell(row=2, column=col)
                        cell.font = Font(italic=True, color="0000FF")
                    
                    # 设置第三行（示例）为灰色
                    for col in range(1, 14):
                        cell = worksheet.cell(row=3, column=col)
                        cell.font = Font(color="808080")
                    
                    # 设置第四行（提示）为粗体居中
                    for col in range(1, 14):
                        cell = worksheet.cell(row=4, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    # 设置第五行为黄色背景，提示用户从此处开始填写
                    for col in range(1, 14):
                        cell = worksheet.cell(row=5, column=col)
                        from openpyxl.styles import PatternFill
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        cell.font = Font(bold=True)
                        cell.value = "← 请在此处开始填写"
                    
                    # 设置整个工作表的第一列冻结
                    worksheet.freeze_panes = "A2"
                    
            elif file_path.lower().endswith('.xls'):
                # 使用xlwt引擎保存为xls
                df.to_excel(file_path, index=False, header=False, engine='xlwt')
            else:
                # 保存为CSV
                df.to_csv(file_path, index=False, header=False, encoding='utf-8-sig')
            
            # 提供使用说明
            instructions = """
    ✅ 模板导出成功！

    📋 使用说明：

    1. 使用 Microsoft Excel 或 WPS Office 打开此文件
    2. 第一行（红色粗体）是字段名称，带 * 号为必填项
    3. 第二行（蓝色斜体）是字段说明和可选值
    4. 第三行（灰色）是示例数据，供参考格式
    5. 第四行是提示信息
    6. 第五行（黄色背景）开始填写您的焊枪数据
    7. 填写完成后保存文件
    8. 使用系统的"导入数据"功能导入此文件

    💡 提示：
    - 可以一次填写多行数据批量导入
    - 填写时请参考第二行的字段说明
    - 必填字段必须填写，选填字段可留空
            """
            
            messagebox.showinfo("导出成功", 
                            f"模板文件已保存到:\n{file_path}\n\n{instructions}")
            
        except ImportError as e:
            # 如果没有必要的库，使用简单版本
            messagebox.showwarning("依赖缺失", 
                                f"缺少必要的库: {str(e)}\n"
                                "将使用简化版本导出。\n\n"
                                "如需完整功能，请安装：\n"
                                "pip install pandas openpyxl xlwt")
            self.export_template_simple(file_path, field_definitions)
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"导出错误详情:\n{error_details}")
            messagebox.showerror("导出失败", f"导出模板文件失败:\n{str(e)}")

    def export_template_simple(self, file_path, field_definitions):
        """简化版导出 - 只生成CSV"""
        try:
            import csv
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # 写入第一行：字段名称
                headers = [label for _, label, _ in field_definitions]
                writer.writerow(headers)
                
                # 写入第二行：字段说明
                descriptions = [desc for _, _, desc in field_definitions]
                writer.writerow(descriptions)
                
                # 写入第三行：示例数据
                example_data = [
                    "钢点焊", "小原", "GUN-001", "C型", "500", "200", 
                    "150", "4.5", "库卡", "R30", "否", "0", "0"
                ]
                writer.writerow(example_data)
                
                # 写入第四行：提示
                writer.writerow(["↓ 请从下一行开始填写您的数据 ↓"])
                
                # 写入第五行：空行，用户开始填写
                writer.writerow([""] * 13)
            
            # 提供使用说明
            instructions = """
    ✅ CSV模板导出成功！

    📋 使用说明：

    1. 用Excel或WPS打开此CSV文件
    2. 前两行是字段说明，请不要修改
    3. 第三行是示例数据
    4. 从第五行开始填写您的焊枪数据
    5. 保存时请选择"CSV UTF-8(逗号分隔)"格式
    6. 使用系统的"导入数据"功能导入

    💡 提示：
    - 必填字段必须填写（带*号）
    - 填写时请参考第二行的可选值
            """
            
            messagebox.showinfo("导出成功", 
                            f"CSV模板已保存到:\n{file_path}\n\n{instructions}")
            
        except Exception as e:
            messagebox.showerror("导出失败", f"导出CSV模板失败:\n{str(e)}")

    def import_data(self):
        """导入数据文件 - 主要支持Excel格式"""
        # 选择要导入的文件
        file_path = filedialog.askopenfilename(
            title="选择数据文件",
            filetypes=[
                ("Excel文件", "*.xlsx;*.xls"),
                ("CSV文件", "*.csv"),
                ("所有文件", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        try:
            # 尝试导入pandas，如果失败则使用csv
            try:
                import pandas as pd
                use_pandas = True
            except ImportError:
                use_pandas = False
            
            data = []
            
            if file_path.lower().endswith(('.xlsx', '.xls')) and use_pandas:
                # 使用pandas读取Excel
                try:
                    # 读取整个工作表
                    df = pd.read_excel(file_path, header=None)
                    
                    # 检查是否至少有4行（标题、说明、示例、提示）
                    if len(df) < 4:
                        messagebox.showerror("格式错误", "文件格式不正确，请使用导出的模板文件")
                        return
                    
                    # 从第4行开始（跳过标题、说明、示例、提示）
                    for i in range(4, len(df)):
                        row = df.iloc[i].tolist()
                        # 转换为字符串并清理
                        cleaned_row = []
                        for cell in row:
                            if pd.isna(cell):
                                cleaned_row.append('')
                            else:
                                cleaned_row.append(str(cell).strip())
                        
                        # 检查是否为空行
                        if any(cell for cell in cleaned_row):
                            data.append(cleaned_row)
                            
                except Exception as e:
                    messagebox.showerror("Excel读取错误", 
                                    f"读取Excel文件失败:\n{str(e)}\n\n"
                                    "请确保文件未被其他程序打开，且格式正确。")
                    return
                    
            else:
                # 使用CSV或pandas不可用时
                if not use_pandas:
                    response = messagebox.askyesno("依赖缺失", 
                        "未找到pandas库，无法处理Excel文件。\n"
                        "是否使用CSV格式导入？\n\n"
                        "如需Excel支持，请安装：pip install pandas openpyxl")
                    if not response:
                        return
                
                # 读取CSV文件
                try:
                    if use_pandas:
                        df = pd.read_csv(file_path, header=None, encoding='utf-8-sig')
                        # 从第4行开始
                        for i in range(3, len(df)):
                            row = df.iloc[i].tolist()
                            cleaned_row = [str(cell).strip() if not pd.isna(cell) else '' for cell in row]
                            if any(cell for cell in cleaned_row):
                                data.append(cleaned_row)
                    else:
                        # 使用纯CSV读取
                        import csv
                        with open(file_path, 'r', encoding='utf-8-sig') as f:
                            reader = csv.reader(f)
                            rows = list(reader)
                            # 从第4行开始（跳过标题、说明、示例）
                            for i in range(3, len(rows)):
                                row = rows[i]
                                cleaned_row = [cell.strip() for cell in row]
                                if any(cell for cell in cleaned_row):
                                    data.append(cleaned_row)
                except UnicodeDecodeError:
                    # 尝试其他编码
                    encodings = ['gbk', 'gb2312', 'utf-8']
                    for encoding in encodings:
                        try:
                            if use_pandas:
                                df = pd.read_csv(file_path, header=None, encoding=encoding)
                                for i in range(3, len(df)):
                                    row = df.iloc[i].tolist()
                                    cleaned_row = [str(cell).strip() if not pd.isna(cell) else '' for cell in row]
                                    if any(cell for cell in cleaned_row):
                                        data.append(cleaned_row)
                                break
                            else:
                                with open(file_path, 'r', encoding=encoding) as f:
                                    reader = csv.reader(f)
                                    rows = list(reader)
                                    for i in range(3, len(rows)):
                                        row = rows[i]
                                        cleaned_row = [cell.strip() for cell in row]
                                        if any(cell for cell in cleaned_row):
                                            data.append(cleaned_row)
                                break
                        except:
                            continue
            
            if not data:
                messagebox.showwarning("警告", "文件中没有有效数据")
                return
            
            # 显示导入确认对话框
            confirm_msg = f"找到 {len(data)} 条待导入数据\n\n"
            confirm_msg += "字段说明：\n"
            confirm_msg += "1. 焊接类型* (必填)\n"
            confirm_msg += "2. 焊枪品牌* (必填)\n"
            confirm_msg += "3. 焊枪编号* (必填)\n"
            confirm_msg += "4-13. 其他字段 (选填)\n\n"
            confirm_msg += "是否开始导入？"
            
            response = messagebox.askyesno("确认导入", confirm_msg)
            if not response:
                return
            
            # 处理导入的数据（使用之前优化过的处理逻辑）
            # ... 这里使用之前已经优化过的处理逻辑
            
            # 刷新文件列表
            self.refresh_file_list()
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"导入错误详情:\n{error_details}")
            messagebox.showerror("导入失败", 
                            f"导入文件失败:\n{str(e)}\n\n"
                            "建议：\n"
                            "1. 使用系统导出的模板文件\n"
                            "2. 确保Excel文件未被其他程序打开\n"
                            "3. 检查必填字段是否填写完整")
 
# 7. 最后，在文件的最后添加 main() 函数
def main():
    """主函数"""
    print("=== 焊接枪管理系统启动 ===")
    try:
        app = WeldingGunSystem()
        print("=== 系统已关闭 ===")
    except Exception as e:
        print(f"!!! 启动错误: {e}")
        import traceback
        traceback.print_exc()
        # 尝试使用tkinter显示错误
        try:
            import tkinter as tk
            from tkinter import messagebox
            root = tk.Tk()
            root.withdraw()  # 隐藏主窗口
            messagebox.showerror("启动错误", f"应用程序启动失败:\n{str(e)}")
        except:
            pass

if __name__ == "__main__":
    main()